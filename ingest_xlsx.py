#!/usr/bin/env python3
"""
Stream large .xlsx files into a database staging table.

The goal is to avoid putting massive XLSX contents into an LLM context.
The script stores:
  - source file metadata in source_files
  - raw rows as JSON in raw_records

Examples:
  # Inspect sheet names and sample rows only
  python ingest_xlsx.py \
    --db sqlite:///visa.db \
    --file data/h1b_fy2024.xlsx \
    --inspect \
    --sample-rows 5

  # Ingest one file
  python ingest_xlsx.py \
    --db sqlite:///visa.db \
    --file data/h1b_fy2024.xlsx \
    --url "https://example.gov/h1b_fy2024.xlsx" \
    --publisher "USCIS" \
    --downloaded-at "2026-06-16" \
    --replace

  # Ingest many files from manifest.json
  python ingest_xlsx.py \
    --db sqlite:///visa.db \
    --manifest manifest.json \
    --replace
"""

import argparse
import hashlib
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import (
    Column,
    DateTime,
    Index,
    Integer,
    MetaData,
    String,
    Table,
    Text,
    create_engine,
    delete,
    select,
    update,
)

metadata = MetaData()

source_files = Table(
    "source_files",
    metadata,
    Column("id", Integer, primary_key=True),
    Column("path", String(1024), unique=True, nullable=False),
    Column("filename", String(512)),
    Column("source_url", String(2048)),
    Column("publisher", String(512)),
    Column("downloaded_at", String(64)),
    Column("sha256", String(64)),
    Column("ingested_at", DateTime),
    Column("status", String(32)),
    Column("rows_ingested", Integer),
)

raw_records = Table(
    "raw_records",
    metadata,
    Column("id", Integer, primary_key=True),
    Column("source_file_id", Integer, nullable=False),
    Column("sheet_name", String(255)),
    Column("row_number", Integer),
    Column("data", Text, nullable=False),
)

Index("ix_raw_records_source_file_id", raw_records.c.source_file_id)


def utcnow():
    return datetime.now(timezone.utc).replace(tzinfo=None)


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def clean_header(value, idx: int) -> str:
    if value is None:
        return f"col_{idx + 1}"

    s = str(value).strip().lower()
    s = re.sub(r"[^a-z0-9_]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s or f"col_{idx + 1}"


def make_headers(header_row):
    seen = set()
    headers = []

    for idx, value in enumerate(header_row):
        base = clean_header(value, idx)
        candidate = base
        counter = 2

        while candidate in seen:
            candidate = f"{base}_{counter}"
            counter += 1

        seen.add(candidate)
        headers.append(candidate)

    return headers


def jsonable_value(value):
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def inspect_file(path: Path, sample_rows: int = 5, sheets=None):
    result = {
        "file": str(path),
        "status": "inspected",
        "sheets": [],
    }

    if not path.exists():
        return {
            "file": str(path),
            "status": "error",
            "error": "file not found",
        }

    sheet_filter = set(sheets or [])

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return {
            "file": str(path),
            "status": "error",
            "error": str(exc),
        }

    for ws in wb.worksheets:
        if sheet_filter and ws.title not in sheet_filter:
            continue

        sheet_info = {
            "sheet": ws.title,
            "sample_rows": [],
        }

        try:
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= sample_rows:
                    break

                sheet_info["sample_rows"].append(
                    [jsonable_value(cell) for cell in row]
                )
        except Exception as exc:
            sheet_info["error"] = str(exc)

        result["sheets"].append(sheet_info)

    wb.close()
    return result


def ingest_file(
    engine,
    item: dict,
    batch_size: int = 1000,
    limit_rows: int = None,
    skip_hash: bool = False,
    sheets=None,
    replace: bool = False,
):
    path = Path(item.get("path", "")).expanduser()

    if not path.exists():
        return {
            "file": str(path),
            "status": "error",
            "error": "file not found",
        }

    url = item.get("url") or item.get("source_url")
    publisher = item.get("publisher")
    downloaded_at = item.get("downloaded_at") or item.get("downloadedAt")

    try:
        sha = None if skip_hash else sha256_file(path)
    except Exception as exc:
        return {
            "file": str(path),
            "status": "error",
            "error": f"hashing failed: {exc}",
        }

    with engine.begin() as conn:
        existing = conn.execute(
            select(source_files.c.id).where(source_files.c.path == str(path))
        ).fetchone()

        if existing:
            source_file_id = existing[0]
            conn.execute(
                update(source_files)
                .where(source_files.c.id == source_file_id)
                .values(
                    source_url=url,
                    publisher=publisher,
                    downloaded_at=downloaded_at,
                    sha256=sha,
                    ingested_at=utcnow(),
                    status="started",
                )
            )
        else:
            result = conn.execute(
                source_files.insert().values(
                    path=str(path),
                    filename=path.name,
                    source_url=url,
                    publisher=publisher,
                    downloaded_at=downloaded_at,
                    sha256=sha,
                    ingested_at=utcnow(),
                    status="started",
                    rows_ingested=0,
                )
            )
            source_file_id = result.inserted_primary_key[0]

    if replace:
        with engine.begin() as conn:
            conn.execute(
                delete(raw_records).where(
                    raw_records.c.source_file_id == source_file_id
                )
            )
            conn.execute(
                update(source_files)
                .where(source_files.c.id == source_file_id)
                .values(rows_ingested=0)
            )

    summary = {
        "file": str(path),
        "source_file_id": source_file_id,
        "status": "started",
        "sheets": [],
        "rows_ingested": 0,
    }

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        with engine.begin() as conn:
            conn.execute(
                update(source_files)
                .where(source_files.c.id == source_file_id)
                .values(status="error")
            )

        summary.update(
            status="error",
            error=str(exc),
        )
        return summary

    total_rows = 0
    sheet_filter = set(sheets or [])

    try:
        for ws in wb.worksheets:
            if sheet_filter and ws.title not in sheet_filter:
                continue

            sheet_rows = 0
            rows_iter = ws.iter_rows(values_only=True)

            try:
                header_row = next(rows_iter)
            except StopIteration:
                summary["sheets"].append(
                    {
                        "sheet": ws.title,
                        "rows_ingested": 0,
                        "note": "empty sheet",
                    }
                )
                continue

            headers = make_headers(header_row)
            batch = []
            row_number = 1

            for row in rows_iter:
                row_number += 1

                if limit_rows is not None and total_rows >= limit_rows:
                    break

                if not row:
                    continue

                if all(cell is None for cell in row):
                    continue

                data = {}

                for i, header in enumerate(headers):
                    value = row[i] if i < len(row) else None
                    data[header] = jsonable_value(value)

                batch.append(
                    {
                        "source_file_id": source_file_id,
                        "sheet_name": ws.title,
                        "row_number": row_number,
                        "data": json.dumps(
                            data,
                            ensure_ascii=False,
                            default=str,
                        ),
                    }
                )

                if len(batch) >= batch_size:
                    with engine.begin() as conn:
                        conn.execute(raw_records.insert(), batch)

                    total_rows += len(batch)
                    sheet_rows += len(batch)
                    batch = []

            if batch:
                with engine.begin() as conn:
                    conn.execute(raw_records.insert(), batch)

                total_rows += len(batch)
                sheet_rows += len(batch)

            summary["sheets"].append(
                {
                    "sheet": ws.title,
                    "rows_ingested": sheet_rows,
                }
            )

            if limit_rows is not None and total_rows >= limit_rows:
                break

        wb.close()

        with engine.begin() as conn:
            conn.execute(
                update(source_files)
                .where(source_files.c.id == source_file_id)
                .values(
                    status="done",
                    rows_ingested=total_rows,
                )
            )

        summary.update(
            status="done",
            rows_ingested=total_rows,
        )

    except Exception as exc:
        try:
            wb.close()
        except Exception:
            pass

        with engine.begin() as conn:
            conn.execute(
                update(source_files)
                .where(source_files.c.id == source_file_id)
                .values(status="error")
            )

        summary.update(
            status="error",
            error=str(exc),
        )

    return summary


def build_items(args):
    if args.manifest:
        manifest_path = Path(args.manifest).expanduser()
        raw = json.loads(manifest_path.read_text(encoding="utf-8"))

        if isinstance(raw, dict):
            raw = raw.get("files", [])

        return raw

    if args.file:
        return [
            {
                "path": args.file,
                "url": args.url,
                "publisher": args.publisher,
                "downloaded_at": args.downloaded_at,
            }
        ]

    return []


def main():
    parser = argparse.ArgumentParser(
        description="Ingest large XLSX files into a DB without loading them into LLM context."
    )

    parser.add_argument(
        "--db",
        default="sqlite:///visa_ingest.db",
        help="SQLAlchemy database URL. Example: sqlite:///visa.db",
    )
    parser.add_argument("--file", help="Path to one XLSX file")
    parser.add_argument("--url", help="Source URL for the file")
    parser.add_argument("--publisher", help="Publisher/source name")
    parser.add_argument("--downloaded-at", dest="downloaded_at", help="Download date")
    parser.add_argument(
        "--manifest",
        help="JSON manifest with many files. Expected key: files",
    )
    parser.add_argument(
        "--sheet",
        action="append",
        help="Only process this sheet name. Can be repeated.",
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=1000,
        help="Insert batch size",
    )
    parser.add_argument(
        "--limit",
        type=int,
        help="Optional total row limit for testing",
    )
    parser.add_argument(
        "--skip-hash",
        action="store_true",
        help="Skip SHA256 hashing. Faster for very large files.",
    )
    parser.add_argument(
        "--inspect",
        action="store_true",
        help="Only inspect sheet names and sample rows. Do not ingest.",
    )
    parser.add_argument(
        "--sample-rows",
        type=int,
        default=5,
        help="Number of sample rows to show when using --inspect",
    )
    parser.add_argument(
        "--replace",
        action="store_true",
        help="Delete previously ingested raw_records for the same source file before loading",
    )

    args = parser.parse_args()

    items = build_items(args)

    if not items:
        parser.error("Provide either --file or --manifest")

    engine = create_engine(args.db)
    metadata.create_all(engine)

    results = []

    if args.inspect:
        for item in items:
            path = Path(item.get("path", "")).expanduser()
            results.append(
                inspect_file(
                    path,
                    sample_rows=args.sample_rows,
                    sheets=args.sheet,
                )
            )
    else:
        for item in items:
            results.append(
                ingest_file(
                    engine,
                    item,
                    batch_size=args.batch_size,
                    limit_rows=args.limit,
                    skip_hash=args.skip_hash,
                    sheets=args.sheet,
                    replace=args.replace,
                )
            )

    print(json.dumps(results, indent=2, default=str))

    if any(result.get("status") == "error" for result in results):
        sys.exit(1)


if __name__ == "__main__":
    main()
